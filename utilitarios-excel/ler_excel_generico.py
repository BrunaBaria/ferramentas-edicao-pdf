import openpyxl
import warnings
from absl import logging
logging.set_verbosity(logging.ERROR)
warnings.filterwarnings("ignore")

file_path_original = r"C:\caminho-da-minha-pasta\nome-do-arquivo.xlsx"

def get_excel_content(*, xlsx_file, table_name):
    """Retorna o conteúdo do Excel"""
    workbook = openpyxl.load_workbook(xlsx_file, keep_vba=True, data_only=True)
    table_data = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Verifica se a tabela existe
        if table_name in sheet.tables:
            table = sheet.tables[table_name]
            
            # Obtém o range da tabela
            table_range = table.ref
            min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(table_range)
            
            # Lê o cabeçalho da tabela (primeira linha)
            headers = [sheet.cell(row=min_row, column=col).value for col in range(min_col, max_col + 1)]
            
            # Lê os dados da tabela (excluindo o cabeçalho)
            table_data = []
            for row in range(min_row + 1, max_row + 1):  # Pula a linha de cabeçalho
                row_data = {
                    headers[col_idx]: sheet.cell(row=row, column=min_col + col_idx).value
                    for col_idx in range(len(headers))
                }
                table_data.append(row_data)
            
            # Retorna o dicionário
            return table_data
    return table_data

for row in get_excel_content(xlsx_file=file_path_original, table_name="nome-da-minha-planilha"):
    print(row)
    break
